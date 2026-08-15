"""
Builds dashboard.xlsx from cases.csv + review_log.csv:
  - Cases sheet: raw case dataset
  - Review Log sheet: raw AI/human review data
  - Summary sheet: counts by issue type, severity, and AI-vs-human agreement,
    plus a bar chart.

Run: python scripts/build_dashboard.py
Writes: dashboard.xlsx (repo root)
"""

import csv
import os
from collections import Counter

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = os.path.join(os.path.dirname(__file__), "..")

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def load_csv(name):
    with open(os.path.join(ROOT, name), newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def write_table(ws, rows, fieldnames):
    ws.append(fieldnames)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row in rows:
        ws.append([row.get(fn, "") for fn in fieldnames])
    for i, fn in enumerate(fieldnames, start=1):
        width = min(max(len(fn), *(len(str(r.get(fn, ""))) for r in rows)) + 2, 60) if rows else len(fn) + 2
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"


def main():
    cases = load_csv("cases.csv")
    reviews = load_csv("review_log.csv")

    wb = Workbook()

    # ---- Cases sheet ----
    ws_cases = wb.active
    ws_cases.title = "Cases"
    write_table(ws_cases, cases, list(cases[0].keys()))

    # ---- Review Log sheet ----
    ws_review = wb.create_sheet("Review Log")
    write_table(ws_review, reviews, list(reviews[0].keys()))

    # ---- Summary sheet ----
    ws_sum = wb.create_sheet("Summary")

    category_counts = Counter(c["category"] for c in cases)
    severity_counts = Counter(c["severity"] for c in cases)
    decision_counts = Counter(r["decision"] for r in reviews)

    total = len(reviews)
    accepted = decision_counts.get("Accepted", 0)
    agreement_rate = round(100 * accepted / total, 1) if total else 0

    row = 1
    ws_sum.cell(row=row, column=1, value="NetSage AI — Dashboard Summary").font = Font(bold=True, size=14)
    row += 2

    # Issue type counts
    ws_sum.cell(row=row, column=1, value="Cases by Issue Type").font = Font(bold=True)
    row += 1
    type_start_row = row
    ws_sum.cell(row=row, column=1, value="Category")
    ws_sum.cell(row=row, column=2, value="Count")
    for cell in ws_sum[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    row += 1
    for cat, count in sorted(category_counts.items()):
        ws_sum.cell(row=row, column=1, value=cat)
        ws_sum.cell(row=row, column=2, value=count)
        row += 1
    type_end_row = row - 1
    row += 1

    # Severity counts
    ws_sum.cell(row=row, column=1, value="Cases by Severity").font = Font(bold=True)
    row += 1
    ws_sum.cell(row=row, column=1, value="Severity")
    ws_sum.cell(row=row, column=2, value="Count")
    for cell in ws_sum[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    row += 1
    for sev in ["HIGH", "MEDIUM", "LOW"]:
        ws_sum.cell(row=row, column=1, value=sev)
        ws_sum.cell(row=row, column=2, value=severity_counts.get(sev, 0))
        row += 1
    row += 1

    # AI vs human agreement
    ws_sum.cell(row=row, column=1, value="AI vs Human Review").font = Font(bold=True)
    row += 1
    ws_sum.cell(row=row, column=1, value="Decision")
    ws_sum.cell(row=row, column=2, value="Count")
    for cell in ws_sum[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    row += 1
    for decision in ["Accepted", "Edited", "Rejected"]:
        ws_sum.cell(row=row, column=1, value=decision)
        ws_sum.cell(row=row, column=2, value=decision_counts.get(decision, 0))
        row += 1
    row += 1
    ws_sum.cell(row=row, column=1, value="AI/Human Agreement Rate").font = Font(bold=True)
    ws_sum.cell(row=row, column=2, value=f"{agreement_rate}%").font = Font(bold=True)
    row += 1
    ws_sum.cell(row=row, column=1, value="Cases needing correction (Edited+Rejected)")
    ws_sum.cell(row=row, column=2, value=decision_counts.get("Edited", 0) + decision_counts.get("Rejected", 0))
    row += 2

    ws_sum.column_dimensions["A"].width = 38
    ws_sum.column_dimensions["B"].width = 12

    # Bar chart: cases by issue type
    chart = BarChart()
    chart.title = "Cases by Issue Type"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Category"
    data = Reference(ws_sum, min_col=2, min_row=type_start_row, max_row=type_end_row)
    cats = Reference(ws_sum, min_col=1, min_row=type_start_row + 1, max_row=type_end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 16
    chart.height = 9
    ws_sum.add_chart(chart, "D2")

    out_path = os.path.join(ROOT, "dashboard.xlsx")
    wb.save(out_path)
    print(f"Wrote dashboard to {os.path.abspath(out_path)}")
    print(f"Categories: {dict(category_counts)}")
    print(f"Severity: {dict(severity_counts)}")
    print(f"Review decisions: {dict(decision_counts)} (agreement rate {agreement_rate}%)")


if __name__ == "__main__":
    main()
